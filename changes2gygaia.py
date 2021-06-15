
from deposit.gui import Registry
import deposit.gui.res
import res

from PySide2 import (QtWidgets, QtCore, QtGui)
from openpyxl import load_workbook
from pathlib import Path
from os import system
import psycopg2
import csv
import sys
import os

STRAT_LOOKUP = {
	"Included by": ["Contemporary with", "Included by"],
	"Same as": ["Equal to", "Same as"],
	"Cut by": ["Earlier than", "Cut by"],
	"Covered by": ["Earlier than", "Covered by"],
	"Abutted by": ["Earlier than", "Abutted by"],
	"Overlaps": ["Non-contemporary", "Overlaps"],
}

class MainWindow(QtWidgets.QMainWindow):
	
	def __init__(self):
		
		QtWidgets.QMainWindow.__init__(self)
		
		self.registry = Registry("Gygaia2Deposit")
		
		self.setWindowTitle("Changes2Gygaia")
		self.setWindowIcon(self.get_icon("c2g_icon.svg"))
		self.setStyleSheet("QPushButton {padding: 5px; min-width: 100px;}")
		
		central_widget = QtWidgets.QWidget(self)
		central_widget.setLayout(QtWidgets.QVBoxLayout())
		self.setCentralWidget(central_widget)
		
		self.setGeometry(100, 100, 800, 600)
		
		self.db_host_edit = QtWidgets.QLineEdit()
		self.db_host_edit.setText(self.get_from_registry("db_host"))
		self.db_port_edit = QtWidgets.QLineEdit()
		self.db_port_edit.setText(self.get_from_registry("db_port", "5432"))
		self.db_name_edit = QtWidgets.QLineEdit()
		self.db_name_edit.setText(self.get_from_registry("db_name"))
		self.db_user_edit = QtWidgets.QLineEdit()
		self.db_user_edit.setText(self.get_from_registry("db_user"))
		self.db_pass_edit = QtWidgets.QLineEdit()
		self.db_pass_edit.setEchoMode(QtWidgets.QLineEdit.Password)
		
		self.backup_path_edit = QtWidgets.QLineEdit()
		self.backup_path_edit.setText(self.get_from_registry("backup_path"))
		
		self.backup_path_button = QtWidgets.QToolButton()
		self.backup_path_button.setIcon(self.get_icon("open.svg"))
		self.backup_path_button.setIconSize(QtCore.QSize(24,24))
		self.backup_path_button.clicked.connect(self.on_backup_path)
		
		self.backup_button = QtWidgets.QPushButton("Backup Database")
		self.backup_button.clicked.connect(self.on_backup)
		self.backup_button.setSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Fixed)
		
		self.restore_button = QtWidgets.QPushButton("Restore Database")
		self.restore_button.clicked.connect(self.on_restore)
		self.restore_button.setSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Fixed)
		
		self.xls_path_edit = QtWidgets.QLineEdit()
		self.xls_path_edit.setText(self.get_from_registry("changes_path"))
		
		self.xls_path_button = QtWidgets.QToolButton()
		self.xls_path_button.setIcon(self.get_icon("open.svg"))
		self.xls_path_button.setIconSize(QtCore.QSize(24,24))
		self.xls_path_button.clicked.connect(self.on_xls_path)
		
		self.import_button = QtWidgets.QPushButton("Import Changes")
		self.import_button.clicked.connect(self.on_import)
		self.import_button.setSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Fixed)
		
		self.console_edit = QtWidgets.QTextEdit()
		self.console_edit.setReadOnly(True)
		self.console_edit.setFontPointSize(8)
		
		db_frame = QtWidgets.QGroupBox("Gygaia Database")
		db_frame.setLayout(QtWidgets.QFormLayout())
		db_frame.setSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Fixed)
		db_frame.layout().addRow("Host:", self.db_host_edit)
		db_frame.layout().addRow("Port:", self.db_port_edit)
		db_frame.layout().addRow("Database:", self.db_name_edit)
		db_frame.layout().addRow("Username:", self.db_user_edit)
		db_frame.layout().addRow("Password:", self.db_pass_edit)
		
		backup_frame = QtWidgets.QGroupBox("Backup File")
		backup_frame.setLayout(QtWidgets.QHBoxLayout())
		backup_frame.setSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Fixed)
		backup_frame.layout().addWidget(self.backup_path_edit)
		backup_frame.layout().addWidget(self.backup_path_button)
		
		restore_frame = QtWidgets.QFrame()
		restore_frame.setLayout(QtWidgets.QHBoxLayout())
		restore_frame.layout().setContentsMargins(0, 0, 0, 0)
		restore_frame.layout().addStretch()
		restore_frame.layout().addWidget(self.backup_button)
		restore_frame.layout().addWidget(self.restore_button)
		restore_frame.layout().addStretch()
		
		xls_frame = QtWidgets.QGroupBox("Changes File")
		xls_frame.setLayout(QtWidgets.QHBoxLayout())
		xls_frame.setSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Fixed)
		xls_frame.layout().addWidget(self.xls_path_edit)
		xls_frame.layout().addWidget(self.xls_path_button)
		
		import_frame = QtWidgets.QFrame()
		import_frame.setLayout(QtWidgets.QHBoxLayout())
		import_frame.layout().setContentsMargins(0, 0, 0, 0)
		import_frame.layout().addStretch()
		import_frame.layout().addWidget(self.import_button)
		import_frame.layout().addStretch()
		
		central_widget.layout().addWidget(db_frame)
		central_widget.layout().addWidget(backup_frame)
		central_widget.layout().addWidget(restore_frame)
		central_widget.layout().addWidget(xls_frame)
		central_widget.layout().addWidget(import_frame)
		central_widget.layout().addWidget(self.console_edit)
	
	def get_icon(self, name):

		path = os.path.join(os.path.dirname(res.__file__), name)
		if os.path.isfile(path):
			return QtGui.QIcon(path)
		path = os.path.join(os.path.dirname(deposit.gui.res.__file__), name)
		if os.path.isfile(path):
			return QtGui.QIcon(path)
		raise Exception("Could not load icon", name)
	
	def get_from_registry(self, key, default = ""):
		
		value = self.registry.get(key)
		if not value:
			value = default
		return value
	
	def write_console(self, text):
		
		text = str(text)
		if not text:
			return
		self.console_edit.append(text)
		QtWidgets.QApplication.processEvents()
	
	def load_xlsx(self, path):
		
		if not os.path.isfile(path):
			return []
		wb = load_workbook(filename = path, read_only = True)
		ws = None
		for sheet in wb.sheetnames:
			ws = wb[sheet]
			break
		if ws is None:
			return []
		changes = []
		for row in ws.iter_rows(min_row = 2):
			action, feature_src, relation, feature_tgt = [str(val.value).strip() for val in row[:4]]
			changes.append([action, feature_src, relation, feature_tgt])
		return changes
	
	def load_csv(self, path):
		
		if not os.path.isfile(path):
			return []
		changes = []
		with open(path, "r", newline = "") as f:
			reader = csv.reader(f, dialect = csv.excel, quoting=csv.QUOTE_ALL)
			next(reader)
			for row in reader:
				action, feature_src, relation, feature_tgt = [str(val).strip() for val in row[:4]]
				changes.append([action, feature_src, relation, feature_tgt])
		return changes
	
	def backup_db(self, path, cursor):
		
		def format_value(val):
			
			if val is None:
				return "NULL"
			elif isinstance(val, str):
				return "'%s'" % (val)
			return str(val)
		
		cursor.execute("SELECT * FROM excavation.contexts_stratigraphic_relationships")
		column_names = []
		columns_descr = cursor.description
		for c in columns_descr:
			column_names.append(c[0])
		insert_prefix = 'INSERT INTO excavation.contexts_stratigraphic_relationships (%s) VALUES ' % (','.join(column_names))
		rows = cursor.fetchall()
		text = ""
		for row in rows:
			text += "%s (%s);\n" % (insert_prefix, ', '.join([format_value(rd) for rd in row]))
		with open(path, "w") as f:
			f.write(text)
		
		return len(rows)
	
	@QtCore.Slot()
	def on_backup_path(self):
		
		name = self.db_name_edit.text()
		if not name:
			name = "gygaia_strat_backup"
		default_path = self.registry.get("backup_path")
		if default_path:
			default_path = os.path.split(default_path)[0]
		if not default_path:
			default_path = os.path.join(str(Path.home()), "Desktop", name)
		path, _ = QtWidgets.QFileDialog.getSaveFileName(None, "Select Backup File", default_path, "SQL File (*.sql)")
		if path:
			self.backup_path_edit.setText(path)
	
	@QtCore.Slot()
	def on_xls_path(self):
		
		default_path = self.registry.get("changes_path")
		if default_path:
			default_path = os.path.split(default_path)[0]
		if not default_path:
			default_path = os.path.join(str(Path.home()), "Desktop")
		path, _ = QtWidgets.QFileDialog.getOpenFileName(None, caption = "Select Changes File", dir = default_path, filter = "Excel 2007+ Workbook (*.xlsx);;Comma-separated Values (*.csv)")
		if path:
			self.xls_path_edit.setText(path)
	
	@QtCore.Slot()
	def on_backup(self):
		
		db_host = self.db_host_edit.text().strip()
		db_port = self.db_port_edit.text().strip()
		db_name = self.db_name_edit.text().strip()
		db_user = self.db_user_edit.text().strip()
		db_pass = self.db_pass_edit.text().strip()
		backup_path = self.backup_path_edit.text().strip()
		if "" in [db_host, db_port, db_name, db_user, db_pass, backup_path]:
			self.write_console("Backup Failed: Please fill in all fields")
			return
		try:
			conn = psycopg2.connect(dbname = db_name, user = db_user, password = db_pass, host = db_host, port = int(db_port))
			cursor = conn.cursor()
		except:
			self.write_console("Backup Failed: Could not connect to Gygaia database")
			return
		
		self.registry.set("db_host", db_host)
		self.registry.set("db_port", str(db_port))
		self.registry.set("db_name", db_name)
		self.registry.set("db_user", db_user)
		self.registry.set("backup_path", backup_path)
		
		n_rows = self.backup_db(backup_path, cursor)
		
		self.write_console("Backup succesfull. Backed up %d rows" % (n_rows))
	
	@QtCore.Slot()
	def on_restore(self):
		
		db_host = self.db_host_edit.text().strip()
		db_port = self.db_port_edit.text().strip()
		db_name = self.db_name_edit.text().strip()
		db_user = self.db_user_edit.text().strip()
		db_pass = self.db_pass_edit.text().strip()
		backup_path = self.backup_path_edit.text().strip()
		if "" in [db_host, db_port, db_name, db_user, db_pass, backup_path]:
			self.write_console("Restore Failed: Please fill in all fields")
			return
		if not os.path.isfile(backup_path):
			self.write_console("Restore Failed: Backup file not found")
			return
		try:
			conn = psycopg2.connect(dbname = db_name, user = db_user, password = db_pass, host = db_host, port = int(db_port))
			cursor = conn.cursor()
		except:
			self.write_console("Restore Failed: Could not connect to Gygaia database")
			return
		try:
			with open(backup_path, "r") as f:
				query = f.read()
		except:
			self.write_console("Restore Failed: Could not read backup file")
			return
		
		self.registry.set("db_host", db_host)
		self.registry.set("db_port", str(db_port))
		self.registry.set("db_name", db_name)
		self.registry.set("db_user", db_user)
		self.registry.set("backup_path", backup_path)
		
		cursor.execute("DELETE FROM excavation.contexts_stratigraphic_relationships")
		conn.commit()
		cursor.execute(query)
		conn.commit()
		
		cursor.execute("SELECT * FROM excavation.contexts_stratigraphic_relationships")
		n_restored = len(cursor.fetchall())
		cursor.close()
		conn.close()
		
		self.write_console("Restore succesfull. Restored %d rows" % (n_restored))
	
	@QtCore.Slot()
	def on_import(self):
		
		def add_relation(area_easting_src, area_northing_src, context_number_src, area_easting_tgt, area_northing_tgt, context_number_tgt, rel, subtype):
			
			query = "INSERT INTO excavation.contexts_stratigraphic_relationships (area_easting, area_northing, context_number, related_area_easting, related_area_northing, related_context_number, stratigraphic_relationship_type, stratigraphic_relationship_subtype) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)"
			try:
				cursor.execute(query, (area_easting_src, area_northing_src, context_number_src, area_easting_tgt, area_northing_tgt, context_number_tgt, rel, subtype))
				conn.commit()
				return True
			except:
				_, exc_value, _ = sys.exc_info()
				self.write_console("Database Insert Failed")
				self.write_console(str(exc_value))
				return False
		
		def remove_relation(area_easting_src, area_northing_src, context_number_src, area_easting_tgt, area_northing_tgt, context_number_tgt, rel, subtype):
			query = "DELETE FROM excavation.contexts_stratigraphic_relationships WHERE (area_easting = %s) AND (area_northing = %s) AND (context_number = %s) AND (related_area_easting = %s) AND (related_area_northing = %s) AND (related_context_number = %s) AND (stratigraphic_relationship_type = %s) AND (stratigraphic_relationship_subtype = %s);"
			try:
				cursor.execute(query, (area_easting_src, area_northing_src, context_number_src, area_easting_tgt, area_northing_tgt, context_number_tgt, rel, subtype))
				conn.commit()
				return True
			except:
				_, exc_value, _ = sys.exc_info()
				self.write_console("Database Deletion Failed")
				self.write_console(str(exc_value))
				return False
		
		db_host = self.db_host_edit.text().strip()
		db_port = self.db_port_edit.text().strip()
		db_name = self.db_name_edit.text().strip()
		db_user = self.db_user_edit.text().strip()
		db_pass = self.db_pass_edit.text().strip()
		changes_path = self.xls_path_edit.text().strip()
		backup_path = self.backup_path_edit.text().strip()
		if "" in [db_host, db_port, db_name, db_user, db_pass, changes_path, backup_path]:
			self.write_console("Import Failed: Please fill in all fields")
			return
			
		try:
			conn = psycopg2.connect(dbname = db_name, user = db_user, password = db_pass, host = db_host, port = int(db_port))
			cursor = conn.cursor()
		except:
			self.write_console("Export Failed: Could not connect to Gygaia database")
			return
			
		if not os.path.isfile(changes_path):
			self.write_console("Import Failed: Changes file not found")
			return
		
		n_rows = self.backup_db(backup_path, cursor)
		self.write_console("Backup succesfull. Backed up %d rows" % (n_rows))
		
		self.registry.set("db_host", db_host)
		self.registry.set("db_port", str(db_port))
		self.registry.set("db_name", db_name)
		self.registry.set("db_user", db_user)
		self.registry.set("changes_path", changes_path)
		self.registry.set("backup_path", backup_path)
		
		changes = []
		if changes_path.lower().endswith(".xlsx"):
			changes = self.load_xlsx(changes_path)
		elif changes_path.lower().endswith(".csv"):
			changes = self.load_csv(changes_path)
			
		if not changes:
			self.write_console("Import Failed: No changes found in file")
			return
		
		reply = QtWidgets.QMessageBox.question(self, "Modify Gygaia Database?", "Apply %d changes to Gygaia database?" % (len(changes)), QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No)
		if reply != QtWidgets.QMessageBox.Yes:
			return
		
		n_processed = 0
		for action, feature_src, relation, feature_tgt in changes:
			if relation not in STRAT_LOOKUP:
				self.write_console("Unrecognized relation: %s" % (relation))
				continue
			rel, subtype = STRAT_LOOKUP[relation]
			
			area_easting_src, area_northing_src, context_number_src = feature_src.split(".")
			area_easting_tgt, area_northing_tgt, context_number_tgt = feature_tgt.split(".")
			
			if action == "add_relation":
				if add_relation(area_easting_src, area_northing_src, context_number_src, area_easting_tgt, area_northing_tgt, context_number_tgt, rel, subtype):
					n_processed += 1
				
			elif action == "remove_relation":
				if remove_relation(area_easting_src, area_northing_src, context_number_src, area_easting_tgt, area_northing_tgt, context_number_tgt, rel, subtype):
					n_processed += 1
				
			elif action == "reverse_relation":
				ret = remove_relation(area_easting_src, area_northing_src, context_number_src, area_easting_tgt, area_northing_tgt, context_number_tgt, rel, subtype)
				ret = ret and add_relation(area_easting_tgt, area_northing_tgt, context_number_tgt, area_easting_src, area_northing_src, context_number_src, rel, subtype)
				if ret:
					n_processed += 1
				
			else:
				self.write_console("Unrecognized action: %s" % (action))
		
		self.write_console("Import to Gygaia database successfull. Processed %d changes" % (n_processed))
	
if __name__ == '__main__':
	
	system("title Changes2Gygaia")
	
	app = QtWidgets.QApplication(sys.argv)
	main = MainWindow()
	main.show()
	app.exec_()
	
	